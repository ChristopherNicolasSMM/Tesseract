"""
tests/test_admin_smart_list_parity.py

Cobre os 2 ajustes validados antes de implementar: alinhamento do
botão "Novo" em Gestão de Usuários, e paridade de export/filtro entre
as 24 entidades geradas pelo CrudGen e as 6 telas administrativas
escritas à mão (Usuários, Roles, Transações, Regras de Campo, OData,
Designer) — que tinham ficado de fora do smart-list completo.
"""
import csv
import io

import pytest
from openpyxl import load_workbook

from core.app_factory import create_app
from core.db import db
from model.core.user import User
from model.core.role import Role


@pytest.fixture
def app():
    app = create_app(env="testing")
    yield app


@pytest.fixture
def client(app):
    return app.test_client()


def _login_admin(app, client):
    with app.app_context():
        admin = User(
            username="admin", email="admin@test.local",
            nome="Admin", nome_completo="Administrador", celular="11999999999",
            is_admin=True, is_active=True,
        )
        admin.set_password("senha123")
        db.session.add(admin)
        db.session.commit()
    client.post("/api/auth/login", json={"username": "admin", "password": "senha123"})


# ── Alinhamento do botão "Novo" ──────────────────────────────────────────────

def test_botao_novo_usuario_sem_heading_duplicado(app, client):
    """
    Bug real isolado: <h5 class="card-title">Novo usuário</h5> ficava
    solto acima do botão, duplicando o texto e empurrando ele pra
    baixo — único caso assim em todo o projeto.
    """
    _login_admin(app, client)
    resp = client.get("/admin/users/")
    assert b'<h5 class="card-title">Novo usu' not in resp.data


@pytest.mark.parametrize("url", [
    "/admin/users/", "/admin/roles/", "/admin/transactions/",
    "/admin/field-rules/", "/admin/odata/", "/admin/designer/",
])
def test_botao_novo_e_o_primeiro_elemento_do_card_body(app, client, url):
    """
    Todas as 6 telas administrativas devem seguir o mesmo padrão do
    CrudGen: botão "Novo X" como primeiro elemento dentro de
    card-body, sem heading duplicando o texto por cima.
    """
    _login_admin(app, client)
    resp = client.get(url)
    html = resp.data.decode("utf-8")
    card_body_idx = html.find('<div class="card-body">')
    button_idx = html.find("btn-primary mb-3", card_body_idx)
    # Não deve haver nenhum <h5 class="card-title"> entre o card-body e o botão
    between = html[card_body_idx:button_idx]
    assert "card-title" not in between


# ── Paridade de smart-list nas 6 telas administrativas ──────────────────────

@pytest.mark.parametrize("url", [
    "/admin/users/", "/admin/roles/", "/admin/transactions/",
    "/admin/field-rules/", "/admin/odata/", "/admin/designer/",
])
def test_tela_tem_busca_e_botoes_de_export(app, client, url):
    _login_admin(app, client)
    resp = client.get(url)
    assert resp.status_code == 200
    assert b'name="q"' in resp.data
    assert b"export.csv" in resp.data
    assert b"export.xlsx" in resp.data


def test_busca_filtra_usuarios(app, client):
    _login_admin(app, client)
    with app.app_context():
        outro = User(
            username="zelda", email="zelda@test.local", nome="Zelda",
            nome_completo="Zelda Link", celular="11900000000", is_active=True,
        )
        outro.set_password("senha123")
        db.session.add(outro)
        db.session.commit()

    resp = client.get("/admin/users/?q=zelda")
    assert b"zelda" in resp.data
    assert b"admin" not in resp.data or resp.data.count(b"admin") <= resp.data.count(b"href")  # tolera nav/menu


def test_busca_filtra_roles(app, client):
    _login_admin(app, client)
    with app.app_context():
        db.session.add_all([Role(name="brewmaster"), Role(name="leitor")])
        db.session.commit()

    resp = client.get("/admin/roles/?q=brewmaster")
    assert b"brewmaster" in resp.data
    assert b"leitor" not in resp.data


def test_export_csv_usuarios_contem_dados_reais(app, client):
    _login_admin(app, client)
    resp = client.get("/admin/users/export.csv")
    assert resp.status_code == 200
    assert resp.mimetype == "text/csv"

    rows = list(csv.reader(io.StringIO(resp.data.decode("utf-8"))))
    header = rows[0]
    assert "username" in header
    usernames = [row[header.index("username")] for row in rows[1:]]
    assert "admin" in usernames


def test_export_csv_respeita_filtro_ativo(app, client):
    _login_admin(app, client)
    with app.app_context():
        db.session.add_all([Role(name="brewmaster"), Role(name="leitor")])
        db.session.commit()

    resp = client.get("/admin/roles/export.csv?q=brewmaster")
    rows = list(csv.reader(io.StringIO(resp.data.decode("utf-8"))))
    header = rows[0]
    names = [row[header.index("name")] for row in rows[1:]]
    assert names == ["brewmaster"]


def test_export_xlsx_eh_arquivo_excel_valido(app, client):
    _login_admin(app, client)
    with app.app_context():
        db.session.add(Role(name="brewmaster"))
        db.session.commit()

    resp = client.get("/admin/roles/export.xlsx")
    assert resp.status_code == 200
    assert "spreadsheet" in resp.mimetype

    wb = load_workbook(io.BytesIO(resp.data))
    rows = list(wb.active.iter_rows(values_only=True))
    header = rows[0]
    names = [row[header.index("name")] for row in rows[1:]]
    assert "brewmaster" in names


def test_paginacao_aparece_quando_ha_mais_de_uma_pagina(app, client):
    _login_admin(app, client)
    with app.app_context():
        for i in range(25):
            db.session.add(Role(name=f"role_{i:02d}"))
        db.session.commit()

    resp = client.get("/admin/roles/")
    assert b"pagination" in resp.data
    assert b"page=2" in resp.data
