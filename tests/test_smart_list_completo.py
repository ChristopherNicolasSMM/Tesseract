"""
tests/test_smart_list_completo.py

Cobre os 3 itens que tinham sido deixados de fora do smart-list-lite:
export (CSV/Excel), configuração de colunas por usuário, e filtros
tipados (boolean + @choices).
"""
import csv
import io

import pytest
from openpyxl import load_workbook

from core.app_factory import create_app
from core.db import db
from model.core.user import User


@pytest.fixture
def app():
    app = create_app(env="testing")
    yield app


@pytest.fixture
def client(app):
    return app.test_client()


def _login_admin(app, client):
    with app.app_context():
        admin = User(
            username="admin", email="admin@test.local",
            nome="Admin", nome_completo="Administrador", celular="11999999999",
            is_admin=True, is_active=True,
        )
        admin.set_password("senha123")
        db.session.add(admin)
        db.session.commit()
    client.post("/api/auth/login", json={"username": "admin", "password": "senha123"})


def _seed_functions(client):
    client.post("/device-manager/device-functions/", data={"name": "temp", "display_name": "Temperatura", "category": "sensor"})
    client.post("/device-manager/device-functions/", data={"name": "press", "display_name": "Pressao", "category": "sensor"})
    client.post("/device-manager/device-functions/", data={"name": "heat", "display_name": "Aquecedor", "category": "actuator"})


# ── Filtro tipado (@choices) ─────────────────────────────────────────────────

def test_filtro_choices_aparece_na_tela(app, client):
    _login_admin(app, client)
    _seed_functions(client)
    resp = client.get("/device-manager/device-functions/")
    assert b"filter_category" in resp.data
    assert b">sensor<" in resp.data
    assert b">actuator<" in resp.data


def test_filtro_choices_filtra_corretamente(app, client):
    _login_admin(app, client)
    _seed_functions(client)

    resp = client.get("/device-manager/device-functions/?filter_category=sensor")
    assert b"temp" in resp.data
    assert b"press" in resp.data
    assert b"heat" not in resp.data

    resp = client.get("/device-manager/device-functions/?filter_category=actuator")
    assert b"heat" in resp.data
    assert b"temp" not in resp.data


def test_filtro_boolean_aparece_na_tela(app, client):
    """DeviceMetadata.is_active é boolean — deve virar select Todos/Sim/Não."""
    _login_admin(app, client)
    resp = client.get("/device-manager/device-metadatas/")
    assert b"filter_is_active" in resp.data


def test_filtro_boolean_filtra_corretamente(app, client):
    _login_admin(app, client)
    client.post("/device-manager/device-metadatas/", data={"name": "Ativo", "is_active": "true"})
    client.post("/device-manager/device-metadatas/", data={"name": "Inativo", "is_active": ""})

    resp = client.get("/device-manager/device-metadatas/?filter_is_active=true")
    assert b"Ativo" in resp.data


# ── Configuração de colunas ──────────────────────────────────────────────────

def test_colunas_padrao_mostra_so_o_resumo(app, client):
    _login_admin(app, client)
    _seed_functions(client)
    resp = client.get("/device-manager/device-functions/")
    # cabeçalho padrão: ID + Display Name (campo de resumo) + Ações
    assert resp.data.count(b"<th>") == 3


def test_salvar_preferencia_de_colunas(app, client):
    _login_admin(app, client)
    _seed_functions(client)

    resp = client.post(
        "/device-manager/device-functions/column-prefs",
        data={"columns": ["display_name", "category", "unit"]},
        follow_redirects=True,
    )
    assert b"Colunas atualizadas" in resp.data

    resp = client.get("/device-manager/device-functions/")
    assert b"<th>Category</th>" in resp.data
    assert b"<th>Unit</th>" in resp.data


def test_preferencia_de_colunas_eh_por_usuario(app, client):
    _login_admin(app, client)
    _seed_functions(client)
    client.post("/device-manager/device-functions/column-prefs", data={"columns": ["category"]})

    with app.app_context():
        other = User(
            username="outro", email="outro@test.local",
            nome="Outro", nome_completo="Outro Usuario", celular="11988888888",
            is_admin=True, is_active=True,
        )
        other.set_password("senha123")
        db.session.add(other)
        db.session.commit()

    client.post("/api/auth/logout")
    client.post("/api/auth/login", json={"username": "outro", "password": "senha123"})

    resp = client.get("/device-manager/device-functions/")
    # outro usuário não configurou nada -> volta ao padrão (campo de resumo)
    assert b"<th>Category</th>" not in resp.data


def test_preferencia_de_colunas_persiste_no_banco(app, client):
    _login_admin(app, client)
    client.post("/device-manager/device-functions/column-prefs", data={"columns": ["category", "unit"]})

    with app.app_context():
        from model.core.user_list_preference import UserListPreference
        pref = UserListPreference.query.filter_by(list_key="device_functions").first()
        assert pref is not None
        assert set(pref.visible_columns_json) == {"category", "unit"}


# ── Export ───────────────────────────────────────────────────────────────────

def test_export_csv_contem_os_dados(app, client):
    _login_admin(app, client)
    _seed_functions(client)

    resp = client.get("/device-manager/device-functions/export.csv")
    assert resp.status_code == 200
    assert resp.mimetype == "text/csv"

    rows = list(csv.reader(io.StringIO(resp.data.decode("utf-8"))))
    header = rows[0]
    assert "name" in header
    assert "category" in header

    names = [row[header.index("name")] for row in rows[1:]]
    assert set(names) == {"temp", "press", "heat"}


def test_export_csv_respeita_filtro_ativo(app, client):
    _login_admin(app, client)
    _seed_functions(client)

    resp = client.get("/device-manager/device-functions/export.csv?filter_category=sensor")
    rows = list(csv.reader(io.StringIO(resp.data.decode("utf-8"))))
    header = rows[0]
    names = [row[header.index("name")] for row in rows[1:]]
    assert set(names) == {"temp", "press"}


def test_export_xlsx_eh_um_arquivo_excel_valido(app, client):
    _login_admin(app, client)
    _seed_functions(client)

    resp = client.get("/device-manager/device-functions/export.xlsx")
    assert resp.status_code == 200
    assert "spreadsheet" in resp.mimetype

    wb = load_workbook(io.BytesIO(resp.data))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    assert "name" in header

    names = [row[header.index("name")] for row in rows[1:]]
    assert set(names) == {"temp", "press", "heat"}


def test_export_nao_inclui_registros_na_lixeira(app, client):
    _login_admin(app, client)
    client.post("/device-manager/device-functions/", data={"name": "lixo", "display_name": "Lixo", "category": "sensor"})

    with app.app_context():
        from addons.addon_device_manager.root.model.device_function import DeviceFunction
        item_id = DeviceFunction.query.filter_by(name="lixo").first().id

    client.post(f"/device-manager/device-functions/{item_id}/trash")

    resp = client.get("/device-manager/device-functions/export.csv")
    assert b"lixo" not in resp.data
